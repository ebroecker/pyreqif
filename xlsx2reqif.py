#!/usr/bin/env python3
import os
import zipfile
import xml.etree.ElementTree
import openpyxl
import pyreqif.create
import pyreqif.reqif
import io
import sys

def get_images_from_excel(excel_file, output_file):
    out_zip = zipfile.ZipFile(document_title + ".reqifz", 'w', zipfile.ZIP_DEFLATED)


    in_excel = zipfile.ZipFile(excel_file)

    if "xl/drawings/drawing1.xml" not in in_excel.namelist():
        return []

    ns = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
    a_ns = "{http://schemas.openxmlformats.org/drawingml/2006/main}"

    drawing_source = in_excel.open("xl/drawings/drawing1.xml")
    drawing_tree = xml.etree.ElementTree.parse(drawing_source)
    drawing_root = drawing_tree.getroot()

    images = []
    for anchor in drawing_root:
        row = None
        col = None
        img_ref = None

        row = anchor.find(ns + "from/" + ns + "row").text
        col = anchor.find(ns + "from/" + ns + "col").text
        img_ref = anchor.find(ns + "pic/" + ns + "blipFill/" + a_ns + "blip").attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"]

        images.append({"row": int(row), "col": int(col), "img_ref" : img_ref})
    drawing_source.close()
    drawing_links_source = in_excel.open("xl/drawings/_rels/drawing1.xml.rels")
    drawing_links_tree = xml.etree.ElementTree.parse(drawing_links_source)
    drawing_links_root = drawing_links_tree.getroot()

    for relation_ship in drawing_links_root:
        id = relation_ship.attrib["Id"]
        target = relation_ship.attrib["Target"]
        for item in images:
            if item["img_ref"] == id:
                item["target"] = target.replace("../", "")
                # copy all images to target:
                out_zip.writestr(target.replace("../", ""), in_excel.read(target.replace("../","xl/")), zipfile.ZIP_DEFLATED)
    drawing_links_source.close()
    out_zip.close()
    return images

def get_images(images, row, col):
    return_pictures = []
    for image in images:
        if image["row"] == row and image["col"] == col:
            return_pictures.append(image)
    return return_pictures




file_name = sys.argv[1]

document_title, _ = os.path.splitext(os.path.basename(file_name))
output_file = document_title + ".reqifz"

images = get_images_from_excel(file_name, output_file)
wb = openpyxl.load_workbook(file_name)
ws = wb.active

columns = []
for col_nr in range(1, ws.max_column+1):
    columns.append(ws.cell(1,col_nr).value)

document_reqif_id = "_{}ReqifId-Header".format(document_title)
spec_reqif_id = "_{}ReqifId--spec".format(document_title)
doc_type_ref = "_doc_type_ref"

#
# create doc:
#
mydoc = pyreqif.create.createDocument(document_reqif_id, title=document_title)
pyreqif.create.addDocType(doc_type_ref, mydoc)
# create primitive datatype
#
pyreqif.create.addDatatype("_datatype_ID", mydoc)

#
# create columns
#
for col in columns:
    pyreqif.create.addReqType("_some_requirement_type_id", "requirement Type", "_reqtype_for_" + col.replace(" ", "_"), col, "_datatype_ID", mydoc)

#
# create document hierarchy head
#
myHierarch = pyreqif.create.createHierarchHead(document_title, typeRef=doc_type_ref, id=spec_reqif_id)

#
# create child elements
#
hierarch_stack = []
last_hierarch_element = myHierarch
for row_nr in range(2,ws.max_row):
    xls_req = dict(zip(columns, [ws.cell(row_nr,x).value for x in range(1,ws.max_column+1)]))
    if not "reqifId" in xls_req:
        xls_req["reqifId"] = pyreqif.create.creatUUID()
    for col in columns:
        # do images:
        pictures = get_images(images, row_nr-1, columns.index(col))
        if type(xls_req[col]) == str:
            xls_req[col] = xls_req[col].replace("<", "&gt;")
            xls_req[col].replace("<", "&lt;")
        if len(pictures) > 0:
            for pic in pictures:
                xls_req[col] = "" if xls_req[col] is None else xls_req[col]
                xls_req[col] += "<img src={}>".format(pic["target"])
        if xls_req[col] is not None:
            pyreqif.create.addReq(xls_req["reqifId"],"_some_requirement_type_id", "<div>" + str(xls_req[col]) + "</div>", "_reqtype_for_" + col.replace(" ","_"), mydoc)



    # do hierarchy
    hierarch_element = pyreqif.create.createHierarchElement(xls_req["reqifId"])
    level = ws.row_dimensions[row_nr].outline_level+1
    if level > len(hierarch_stack):
        hierarch_stack.append(last_hierarch_element)
    elif level < len(hierarch_stack):
        hierarch_stack = hierarch_stack[0:level]

    current_head = hierarch_stack[level - 1]
    current_head.addChild(hierarch_element)
    last_hierarch_element = hierarch_element

mydoc.hierarchy.append(myHierarch)


#
# safe reqif to string
strIO = io.StringIO()
pyreqif.reqif.dump(mydoc, strIO)

#add reqifz as zip-file
zip = zipfile.ZipFile(document_title + ".reqifz", 'a', zipfile.ZIP_DEFLATED)
zip.writestr(document_title+'.reqif', strIO.getvalue())

zip.close()




